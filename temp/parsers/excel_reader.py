from openpyxl import load_workbook

wb = load_workbook(filename='1ТИЖДЕНЬ_ІІ_семестр.xlsx')
sheet = wb['1 курс']
for row in sheet.iter_rows():
    print(row)